import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from pathlib import Path

# 获取当前工作目录
base_dir = Path(__file__).parent

# 目标文件
target_file = base_dir / "fc.xlsx"

# 创建或加载目标工作簿
if target_file.exists():
    wb_target = load_workbook(target_file)
else:
    wb_target = openpyxl.Workbook()
    # 删除默认的sheet
    if "Sheet" in wb_target.sheetnames:
        wb_target.remove(wb_target["Sheet"])

# 获取或创建目标工作表（使用第一个sheet，如果不存在则创建）
if len(wb_target.sheetnames) > 0:
    ws_target = wb_target.active
else:
    ws_target = wb_target.create_sheet("Sheet1")

# 遍历所有包含"外周血"的文件夹
row_num = 1  # 从第1行开始写入数据

for folder in base_dir.iterdir():
    if folder.is_dir() and "外周血" in folder.name:
        print(f"处理文件夹: {folder.name}")
        
        # 从文件夹名称提取样本号（去掉" 外周血"后缀）
        sample_id = folder.name.replace(" 外周血", "").strip()
        print(f"  样本号: {sample_id}")
        
        # 查找导入模板文件
        template_files = list(folder.glob("*导入模板.xlsx"))
        
        if not template_files:
            print(f"  未找到导入模板文件")
            continue
        
        template_file = template_files[0]
        print(f"  找到模板文件: {template_file.name}")
        
        try:
            # 打开模板文件
            wb_template = load_workbook(template_file, data_only=True)
            
            # 检查sheet名称
            print(f"  Sheet列表: {wb_template.sheetnames}")
            
            # 明确查找名为"Sheet2"的sheet（注意大小写）
            ws_template = None
            if "Sheet2" in wb_template.sheetnames:
                ws_template = wb_template["Sheet2"]
                print(f"  使用Sheet: Sheet2")
            elif len(wb_template.sheetnames) >= 2:
                # 如果没有找到"Sheet2"，使用第二个sheet（索引为1）
                ws_template = wb_template.worksheets[1]
                print(f"  警告: 未找到Sheet2，使用第二个Sheet: {wb_template.sheetnames[1]}")
            else:
                print(f"  警告: {template_file.name} 没有Sheet2，跳过")
                wb_template.close()
                continue
            
            # 先检查一下C5的值，确认读取位置正确
            test_cell = ws_template.cell(row=5, column=column_index_from_string('C')).value
            test_d_cell = ws_template.cell(row=5, column=column_index_from_string('D')).value
            test_e_cell = ws_template.cell(row=5, column=column_index_from_string('E')).value
            print(f"  验证C5={test_cell}, D5={test_d_cell}, E5={test_e_cell}")
            
            # 读取C5:BK5范围的数据
            source_row = 5
            source_start_col = column_index_from_string('C')  # C列
            source_end_col = column_index_from_string('BK')   # BK列
            
            # 读取数据
            row_data = []
            for col in range(source_start_col, source_end_col + 1):
                cell_value = ws_template.cell(row=source_row, column=col).value
                row_data.append(cell_value)
            
            # 打印前几个值用于调试
            print(f"  读取的前5个值: {row_data[:5]}")
            
            # 写入样本号到A列
            ws_target.cell(row=row_num, column=1, value=sample_id)
            
            # 写入目标文件（从C列开始写入，保持列位置对应）
            for col_idx, value in enumerate(row_data, start=source_start_col):
                ws_target.cell(row=row_num, column=col_idx, value=value)
            
            print(f"  成功复制样本号 {sample_id} 和 {len(row_data)} 个单元格的数据到第 {row_num} 行")
            row_num += 1
            
            wb_template.close()
            
        except Exception as e:
            print(f"  错误: 处理 {template_file.name} 时出错 - {str(e)}")
            continue

# 保存目标文件
wb_target.save(target_file)
print(f"\n所有数据已保存到 {target_file}")

